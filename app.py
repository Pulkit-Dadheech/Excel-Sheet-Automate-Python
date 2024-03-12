import openpyxl as xl
from openpyxl.chart import BarChart, Reference

wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']
cell = sheet.cell(row=1, column=1)

for row in range(2, sheet.max_row + 1):
    try:
        third_cell_value = float(sheet.cell(row=row, column=3).value)
        discounted_value = third_cell_value * 0.9
        sheet.cell(row=row, column=4).value = round(discounted_value, 2)
    except (ValueError, TypeError):
        continue

values = Reference(sheet,min_row=2, max_row=sheet.max_row,min_col=4,max_col=4)
chart = BarChart()
chart.add_data(values)
sheet.add_chart(chart, "E2")

wb.save('transactions.xlsx')
